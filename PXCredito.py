import pyautogui as p
from time import sleep
from openpyxl import Workbook
from openpyxl import load_workbook
import pyperclip as c

index = 1

for i in range(1,4):
    file = r'C:\Users\supor\Desktop\PXTeste.xlsx'
    wb = load_workbook(file)
    ws = wb.active
    num_processo = ws.cell(row=index+1,column=1).value

    #Abrindo o chrome
    p.hotkey('win', 'r')
    p.write('chrome.exe')
    p.press('enter')
    sleep(2)
    #Digitando o link
    p.write('https://pje.trt15.jus.br/consultaprocessual/')
    p.press('enter')
    sleep(3)
    #Copiando o processo
    p.write(num_processo)
    p.press('enter')
    sleep(3)
    #Clicando no 1Grau
    p.moveTo(x=675,y=334)
    sleep(3)
    p.click()
    p.press('enter')
    #Tempo para escrever a captcha
    sleep(8)
    p.press('enter')
    p.moveTo(x=50,y=145)
    sleep(3)
    p.click()
    sleep(4)

    #Copiando o Orgao e colando no excel
    p.tripleClick(x=180, y=222)
    p.hotkey('ctrl','c')
    x = c.paste()
    ws.cell(column= 2 ,row = index+1,value=x)


    #Copiando Autuado
    p.tripleClick(x=180, y=320)
    p.hotkey('ctrl','c')
    x = c.paste()
    ws.cell(column= 3 ,row = index+1,value=x)

    #Copiando Valor
    p.tripleClick(x=180, y=375)
    p.hotkey('ctrl','c')
    x = c.paste()
    ws.cell(column= 4,row = index + 1,value=x)

    #Copiando Justiça
    p.tripleClick(x=180, y=408)
    p.hotkey('ctrl','c')
    x = c.paste()
    ws.cell(column= 5,row = index + 1,value=x)

    # Copiando Polo Ativo
    p.tripleClick(x=542, y=256)
    p.hotkey('ctrl', 'c')
    x = c.paste()
    ws.cell(column= 6, row = index + 1, value=x)

    # Copiando Polo Passivo
    p.tripleClick(x=892, y=258)
    p.hotkey('ctrl', 'c')
    x = c.paste()
    ws.cell(column= 7, row = index + 1, value=x)

    sleep(2)

    #Adiciona +1 no loop para ir para proxima linha de processo
    index += 1
    p.hotkey('alt','f4')
    #Salva cada volta do loop na planilha
    wb.save(file)




