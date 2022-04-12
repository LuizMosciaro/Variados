from distutils.command.config import config
from time import sleep
from pyparsing import Empty
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.relative_locator import locate_with
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from twocaptcha import TwoCaptcha
from datetime import datetime
import os

index = 2
# Abrindo excel
file = r'C:\Users\abcd\Desktop\PXTeste.xlsx'
wb = load_workbook(file)
ws = wb.active
url = 'https://pje.trt15.jus.br/consultaprocessual/'

#Ajuste de data
def ajuste_data(x):    
    if str(x[3:6]) == 'jan':    
        y = x.replace('jan','jan')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data 
    if str(x[3:6]) == 'fev':    
        y = x.replace('fev','feb')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'mar':    
        y = x.replace('mar','mar')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data 
    if str(x[3:6]) == 'abr':    
        y = x.replace('abr','apr')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'mai':    
        y = x.replace('mai','may')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'jun':    
        y = x.replace('jun','jun')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data 
    if str(x[3:6]) == 'jul':    
        y = x.replace('jul','jul')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data 
    if str(x[3:6]) == 'ago':    
        y = x.replace('ago','aug')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data 
    if str(x[3:6]) == 'set':    
        y = x.replace('set','sep')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'out':    
        y = x.replace('out','oct')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data  
    if str(x[3:6]) == 'nov':    
        y = x.replace('nov','nov')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'dez':    
        y = x.replace('dez','dec')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data

config = {
            'apiKey': 'abcd',
            'defaultTimeout':    60,
            'pollingInterval':   10,
        }
solver = TwoCaptcha(**config)

# TODA A CONSULTA ABAIXO FUNCIONA APENAS PARA O PRIMEIRO GRAU DO PROCESSO, Só verifica o 2 grau se não houver
# o 1 grau. Futuramente deixar o processo mais inteligente.
# ________________________________________________________________________
print(f"\nIniciando automação\nTotal de processos na planilha: {int(len(ws['A']))-1}")
for i in range(1, len(ws['A'])):       

    #Configurações do chrome
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument('--incognito')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--log-level=OFF')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    s=Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s,options=chrome_options)
    driver.maximize_window()
    driver.get(url)
    
    # Pegando os processos
    num_processo = ws.cell(row=index, column=1).value
    print("\nRealizando a consulta no processo: " + str(num_processo))

    # Pegando o número do processo do excel e entrando na pagina
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'label-sistema')))
    driver.find_element(By.XPATH, '//*[@id="nrProcessoInput"]').send_keys(num_processo)
    driver.find_element(By.XPATH, '//*[@id="btnPesquisar"]').send_keys(Keys.ENTER)

    # Clicando no botao do 1Grau
    sleep(1)
    try:
        #Há processos que só tem 1grau, uns tem 2grau
        x = driver.find_element(By.XPATH,'//*[@id="painel-escolha-processo"]/button[1]').click()
        if x == None or Empty:
            driver.find_element(By.XPATH,'//*[@id="painel-escolha-processo"]/button[2]').click()
    except:
        pass #isso serve pra continuar se ele não encontrar nenhum dos dois botões

    y=0
    while y < 1: #Esse loop é para ele repetir a tentativa de escrever a captcha
        try: 
            # Escrevendo a Captcha
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,'//*[@id="btnEnviar"]')))
            sleep(1)
            driver.find_element(By.XPATH,'//*[@id="imagemCaptcha"]').screenshot(r'C:\Users\supor\Desktop\captcha.png')
            #Local onde salvar o print
            imagem = (r'C:\Users\supor\Desktop\captcha.png')

            #Pega a imagem e envia para o servidor 2captcha
            result = solver.normal(imagem) #'Result' trás 2 resultados: 'captchaId' e 'code', queremos apenas o 'code'
            print("Captcha:",result['code'])

            driver.find_element(By.XPATH,'//*[@id="captchaInput"]').send_keys(result['code'])
            driver.find_element(By.XPATH, '//*[@id="btnEnviar"]').click()
            y+=1
        except Exception as e:
           driver.refresh()
           continue

    #Aguarda o emblema do Brasil  no documento para poder prosseguir
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.XPATH, '//*[@id="visualizacao-documento"]/div/mat-card/mat-card-header/div[2]')))
    print('Iniciando coleta de dados do processo...')
    
    try:  
        # Essa condição verifica se há o DOCUMENTO de sentença
        botao_sentenca = driver.find_element(By.XPATH,"//*[contains(text(),'(Sentença) -')]")
        botao_sentenca.click()
        
        #Abaixo aguarda aparecer o elemento que corresponde ao status da sentença e escreve na planilha
        sleep(2)
        try:
            procedente = driver.find_element(By.XPATH,"//*[contains(text(),'procedente')]").text
            data_sentenca = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(botao_sentenca))
            date = datetime.strptime(data_sentenca.text, '%d %b %Y').strftime("%d/%m/%Y")
            data_sentenca_ajustada = str(date)[:11]
            print("Data da sentença: ",data_sentenca_ajustada)
     
            #Tratando a string da sentença
            start = procedente.find('Descrição') + 35
            end = procedente.find('o(s) pedido(s)')
            substring_procedente = (procedente[start:end]).capitalize()
            print('Sentença:',substring_procedente)

            #Escrevendo na planilha
            ws.cell(column=13, row=index, value="Sim")
            ws.cell(column=14, row=index, value=substring_procedente)
            ws.cell(column=15, row=index, value=data_sentenca_ajustada)
            

        except:
            procedente = driver.find_element(By.XPATH,"//*[contains(text(),'procedentes')]").text
            data_sentenca = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(botao_sentenca))
            date = datetime.strptime(data_sentenca.text, '%d %b %Y').strftime("%d/%m/%Y")
            data_sentenca_ajustada = str(date)[:11]
            print("Data da sentença: ",data_sentenca_ajustada)
            
            
            #Tratando a string da sentença
            start = procedente.find('Descrição') + 35
            end = procedente.find('o(s) pedido(s)')
            substring_procedente = (procedente[start:end]).capitalize()
            print('Sentença:',substring_procedente)

            #Escrevendo na planilha
            ws.cell(column=13, row=index, value="Sim")
            ws.cell(column=14, row=index, value=substring_procedente)
            ws.cell(column=15, row=index, value=data_sentenca_ajustada)        

    except NoSuchElementException as e:
        print('Não há Sentença')
        ws.cell(column=13, row=index, value="Não")
        ws.cell(column=14, row=index, value='-')
        ws.cell(column=15, row=index, value='-')

    try:
        #Aqui ocorre o tratamento e identificação do Recurso Ordinario    
        if driver.find_element(By.XPATH,'//*[contains(text(),"Recebido(s) o(s) Recurso Ordi")]'):
            recurso_ordinario = driver.find_element(By.XPATH,'//*[contains(text(),"Recebido(s) o(s) Recurso Ordi")]')
            data_recurso_ordinario = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(recurso_ordinario)).text
            data_ajustada = ajuste_data(data_recurso_ordinario)
            print('Data do Recurso',str(data_ajustada)[:11])
            
            #Tratando a string do recurso
            start = recurso_ordinario.text.find('Recebido(s) o(s) Recurso Ordinário de ') + 38
            end = recurso_ordinario.text.find('sem efeito suspensivo')
            recurso_ordinario_substring = (recurso_ordinario.text[start:end])
            print(recurso_ordinario_substring)

            ws.cell(column=16, row=index, value='Sim')
            ws.cell(column=17, row=index, value=recurso_ordinario_substring)
            ws.cell(column=18, row=index, value=str(data_ajustada)[:11])
    except Exception as e:
            print("Não há Recurso Ordinário")
            ws.cell(column=16, row=index, value='Não')
            ws.cell(column=17, row=index, value='-')
            ws.cell(column=18, row=index, value='-')
            pass
            

    # Coletando Juiz: // A coleta do juiz começa aqui pq é antes do click no cabeçalho
    juiz = driver.find_element(By.XPATH, '//*[@id="cabecalhoVisualizador"]/mat-card-subtitle').text
    j = str(juiz)
    inicio = j.find("por") + 3
    final = j.find("em", inicio)
    substrg = j[inicio:final]
    ws.cell(column=11, row=index, value=substrg)

    # Clicando no cabeçalho para abrir o menu de infos do processo:
    WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="titulo-detalhes"]/h1')))
    driver.find_element(By.XPATH, '//*[@id="titulo-detalhes"]/h1').click()

    # Coletando Orgão:
    orgao = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[1]').text
    ws.cell(column=2, row=index, value=orgao)

    # Coletando Distribuido:
    distribuido = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[2]').text
    ws.cell(column=3, row=index, value=distribuido)

    # Coletando Autuado:
    autuado = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[3]').text
    ws.cell(column=4, row=index, value=autuado)

    # Coletando Valor:
    valor = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[4]').text
    ws.cell(column=5, row=index, value=valor)

    # Coletando Justica:
    justica = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dt[5]').text
    ws.cell(column=6, row=index, value=justica)

    # Coletando Polo Ativo:
    polo_ativo = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[1]/pje-parte-processo/section/ul/li/span[2]').text
    ws.cell(column=7, row=index, value=polo_ativo)

    # Coletando Polo Ativo ADVOGADO:
    polo_ativo_adv = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[1]/pje-parte-processo/section/ul/ul/li/small/span').text
    ws.cell(column=8, row=index, value=polo_ativo_adv)

    # Coletando Polo Passivo:
    polo_passivo = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[2]/pje-parte-processo/section/ul/li/span[2]').text
    ws.cell(column=9, row=index, value=polo_passivo)

    # Coletando Polo Passivo ADVOGADO:
    polo_passivo_adv = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[2]/pje-parte-processo/section/ul/ul/li/small/span').text
    ws.cell(column=10, row=index, value=polo_passivo_adv)

    # Coletando Assuntos:
    sleep(1)
    orgao = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]').text +'...'
    x = str(orgao)  # Garantindo que ta em formato de string
    inicio = x.find("Assunto(s):") + 12  # Encontrando o elemento "Assuntos" e definindo que será o inicio da substring, adicionar +11 exclui o elemento (Assunto(s) tem 11 chars)
    final = x.find("...",inicio)  # Note que na linha 123 adicionei reticencias para delimitar como sendo o final da string orgao
    substring = x[inicio:final]  # Criando a substring, que vem da 'x' a partir do inicio até o final
    ws.cell(column=12, row=index, value=substring)  # Inserindo na planilha

    try:
        # Salvando a planilha com os dados
        wb.save(file)
        print(f'Consulta {index - 1} finalizada..')
        os.remove(imagem)
        driver.quit()

        # Add +1 a variavel global index, para que ela seja '2' no proximo loop e leia a próxima linha
        index += 1
    except PermissionError:
        print("ATENÇÃO: Planilha aberta, fechar a mesma e reiniciar a automação")
        break
    

print("\nPrograma finalizado")
print('Saldo:',solver.balance())



