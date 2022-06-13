from plistlib import InvalidFileException
import PySimpleGUI as sg
import openpyxl
#Lembrar de sempre transformar a planilha do fiabilite de 'xls' para 'xlsx'

sg.theme('Dark Grey 10')
layout = [
    [sg.Text('Verificador de Acordos - Cartões DB',justification='center',font='Arial 15 bold'),sg.Image('db_sized5050.png')],
    [sg.Text('Mostre o Relatório do Fiabilite',justification='center')],
    [sg.Input(key='-fia-'),sg.FileBrowse()],
    [sg.Text('Mostre o Relatório do Banco',justification='center')],
    [sg.Input(key='-fran-'),sg.FileBrowse()],
    [sg.OK('OK'),sg.Cancel('Sair'),sg.Text('OBS: Ambos arquivos no formato "xlsx"',font='Arial 8 italic')]
]

window = sg.Window('Verificador de Acordo - DB',layout=layout,no_titlebar=True)

while True:
    
    event,values = window.read()
    
    fiabilite = str(values['-fia-'])
    francesa = str(values['-fran-'])
    
    wb1 = openpyxl.load_workbook(fiabilite)
    ws1 = wb1.active

    wb2 = openpyxl.load_workbook(francesa)
    ws2 = wb2.active

    try:
        #Duas listas pra pegar os valores e nomes
        lista_nomes_fiabilite = []   
        lista_valor_pago_fiabilite = []

        for i in range(2,ws1.max_row): #iterando entre a linha 2 e o numero max de linhas
            nomes = ws1.cell(row=i, column=3).value #identificando a coluna e linha dos nomes
            valor_pago = float(ws1.cell(row=i, column=14).value) #identificando a coluna e linha dos valores pagos
            
            lista_nomes_fiabilite.append(nomes) #colocando na lista
            lista_valor_pago_fiabilite.append(valor_pago)        
            
        #transformando num dicionario sendo a chave o nome e o valor da chave é o valor pago
        dicio_fiabilite = dict(zip(lista_nomes_fiabilite, lista_valor_pago_fiabilite)) 

        #idem linha 12
        lista_nomes_francesa = []   
        lista_valor_pago_francesa = []

        for c in range(2,ws2.max_row):
            nomes2 = ws2.cell(row=c, column=4).value #idem linha 17
            valor_pago2 = float(ws2.cell(row=c, column=10).value) #idem linha 18
            
            lista_nomes_francesa.append(nomes2) #colocando na lista
            lista_valor_pago_francesa.append(valor_pago2)        
            
        #transformando num dicionario sendo a chave o nome e o valor da chave é o valor pago
        dicio_francesa = dict(zip(lista_nomes_francesa, lista_valor_pago_francesa)) 

        #comparando os dois dicionarios e encontrando as diferenças das chaves, depois guardando isso numa lista (que é o nome)
        x = list(set(dicio_fiabilite).difference(dicio_francesa))

        final = {chave: valor for (chave,valor) in dicio_fiabilite.items() if chave in x}
        for i,(k,v) in enumerate(final.items(),1):            
            sg.Print(f'{i}) Nome: {k}\n   Valor: R${v}\n')
        
        total_nomes = len(final.keys())
        total_valor = round(sum(final.values()),2)    
        sg.Print(f'\nTotal encontrados: {total_nomes} || Total Valores: {total_valor}')
        
    except InvalidFileException:
        sg.Print('Verifique os arquivos escolhidos se não estão trocados')
    
    except TypeError:
        sg.Print('Verifique os arquivos escolhidos se não estão trocados')
    
    if event in ('Sair',sg.WIN_CLOSED):
        break
    
    
window.close()