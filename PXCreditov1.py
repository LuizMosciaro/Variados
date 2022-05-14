from distutils.command.config import config
from time import sleep,time
from types import NoneType
from pyparsing import Empty
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.relative_locator import locate_with
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from openpyxl import load_workbook
from twocaptcha import TwoCaptcha
from datetime import datetime
import os

index = 2
# Abrindo excel
file = r'C:\Users\abcd\Desktop\PXTeste.xlsx'
wb = load_workbook(file)
ws = wb.active

start_time = time() #Temporizador

#Tem que estar em maiúsculo o nome do Advogado do POLO PASSIVO
lista_proibida_adv = ['Nome dos advogados proibidos']

#Ajuste de data
def ajuste_data(x):    
    if str(x[3:6]) == 'jan':    
        y = x.replace('jan','jan')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'fev':    
        y = x.replace('fev','feb')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'mar':    
        y = x.replace('mar','mar')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'abr':    
        y = x.replace('abr','apr')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'mai':    
        y = x.replace('mai','may')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'jun':    
        y = x.replace('jun','jun')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'jul':    
        y = x.replace('jul','jul')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'ago':    
        y = x.replace('ago','aug')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'set':    
        y = x.replace('set','sep')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'out':    
        y = x.replace('out','oct')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data  
    if str(x[3:6]) == 'nov':    
        y = x.replace('nov','nov')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data
    if str(x[3:6]) == 'dez':    
        y = x.replace('dez','dec')
        data = datetime.strptime(y,'%d %b %Y').strftime("%d/%m/%Y")
        return data

config = {
            'apiKey': 'chave api',
            'defaultTimeout':    60,
            'pollingInterval':   10,
        }
solver = TwoCaptcha(**config)

max = 2
while ws.cell(row=max,column=1).value != None:
    ws.cell(row=index,column=1).value
    max += 1

print(f"\nIniciando automação\nTotal de processos na planilha: {max-2}")
while ws.cell(row=index,column=1).value != None:   
    #Pegando os processos
    num_processo = ws.cell(row=index, column=1).value
    if num_processo != None:
        numero_trt = num_processo[18:20].replace('0','')
   

    #Configurações do chrome
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_experimental_option("detach", True)
    chrome_options.add_argument('--incognito')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--log-level=OFF')
    chrome_options.add_argument('--disable-software-rasterizer')
    chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
    s=Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s,options=chrome_options)
    driver.maximize_window()
    driver.get(f'https://pje.trt{numero_trt}.jus.br/consultaprocessual/')
   
    print("\nRealizando a consulta em 1ª Instância no processo: " + str(num_processo))   
    # Pegando o número do processo do excel e entrando na pagina
    WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'label-sistema')))
    driver.find_element(By.XPATH, '//*[@id="nrProcessoInput"]').send_keys(num_processo)
    driver.find_element(By.XPATH, '//*[@id="btnPesquisar"]').send_keys(Keys.ENTER)
    
    # Clicando no botao do 1Grau
    sleep(1)
    try:
        #Há processos que só tem 1grau e 2grau, outros pulam diretamente para a tela do processo
        driver.find_element(By.XPATH,'//*[@id="painel-escolha-processo"]/button[1]').click()

    except:
        pass #isso serve pra continuar se ele não encontrar nenhum dos dois botões

    y=0
    while y < 1: #Esse loop é para ele repetir a tentativa de escrever a captcha
        try:
            # Escrevendo a Captcha
            WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,'//*[@id="btnEnviar"]')))
            sleep(1)
            driver.find_element(By.XPATH,'//*[@id="imagemCaptcha"]').screenshot(r'C:\Users\abcd\Desktop\captcha.png')
            #Local onde salvar o print
            imagem = (r'C:\Users\abcd\Desktop\captcha.png')

            #Pega a imagem e envia para o servidor 2captcha
            result = solver.normal(imagem) #'Result' trás 2 resultados: 'captchaId' e 'code', queremos apenas o 'code'
            print("Captcha:",result['code'])

            driver.find_element(By.XPATH,'//*[@id="captchaInput"]').send_keys(result['code'])
            driver.find_element(By.XPATH, '//*[@id="btnEnviar"]').click()
            y+=1
        
        #Aqui até o continue faz com que atualize a pagina caso dê erro na API da captcha
        except Exception as e:
            print("Erro na Captcha:",e)
            driver.refresh()
            sleep(2)
            pass

    #Aguarda o emblema do Brasil  no documento para poder prosseguir
    WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, '//*[@id="visualizacao-documento"]/div/mat-card/mat-card-header/div[2]')))
    print('Iniciando coleta de dados do processo...')

    # Coletando Juiz: // A coleta do juiz começa aqui pq é antes do click no cabeçalho
    juiz = driver.find_element(By.XPATH, '//*[@id="cabecalhoVisualizador"]/mat-card-subtitle').text
    j = str(juiz)
    inicio = j.find("por") + 3
    final = j.find("em", inicio)
    substrg = j[inicio:final]
    ws.cell(column=11, row=index, value=substrg)
    
    #Reclamante e Reclamado
    try:
        sleep(1)
        try:
            #Coletando Reclamante
            reclamantex = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"RECLAMANTE")]')).text
            reclamante = str(reclamantex)+'./*/'
            inicio = reclamante.find("RECLAMANTE:") + 12
            final = reclamante.find('./*/',inicio)
            substrg_reclamante = reclamante[inicio:final]

            #Coletando Reclamado
            reclamadox = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"RECLAMADO")]')).text
            reclamado = str(reclamadox)+'./*/'
            inicio = reclamado.find("RECLAMADO:") + 11
            final = reclamado.find('./*/',inicio)
            substrg_reclamado = reclamado[inicio:final]

            #Printando e escrevendo no excel
            print("Reclamante:",substrg_reclamante)
            ws.cell(column=7, row=index, value=substrg_reclamante)
            print("Reclamado:",substrg_reclamado)
            ws.cell(column=9, row=index, value=substrg_reclamado)
            

        except NoSuchElementException as e:
            #Coletando Reclamante
            reclamantex = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"AUTOR")]')).text
            reclamante = str(reclamantex)+'./*/'
            inicio = reclamante.find("AUTOR:") + 7
            final = reclamante.find('./*/',inicio)
            substrg_reclamante = reclamante[inicio:final]

            #Coletando Reclamado
            reclamadox = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"RÉU")]')).text
            reclamado = str(reclamadox)+'./*/'
            inicio = reclamado.find("REU:") + 5
            final = reclamado.find('./*/',inicio)
            substrg_reclamado = reclamado[inicio:final]

            #Printando e escrevendo no excel
            print("Reclamante:",substrg_reclamante)
            ws.cell(column=7, row=index, value=substrg_reclamante)
            print("Reclamado:",substrg_reclamado)
            ws.cell(column=9, row=index, value=substrg_reclamado)
            
    except:
        print("Reclamante e Reclamado serão captados em outro local")
        wb.save(file)
        pass
    
    # Essa condição verifica se há o DOCUMENTO de sentença
    try:  
        botao_sentenca = driver.find_element(By.XPATH,"//*[contains(text(),'(Sentença)')]")
        botao_sentenca.click()
       
        #Abaixo aguarda aparecer o elemento que corresponde ao status da sentença e escreve na planilha
        sleep(2)
        procedente = driver.find_element(By.XPATH,"//*[contains(text(),'procedente') or contains(text(),'procedente(s)')]").text
        data_sentenca = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(botao_sentenca)).text
        data_ajustada = ajuste_data(data_sentenca)
        print("Data da sentença: ",data_ajustada[:11])
                
        #Tratando a string da sentença
        start = procedente.find('Descrição') + 35
        end = procedente.find('o(s) pedido(s)')
        substring_procedente = (procedente[start:end]).capitalize()
        print('Sentença:',substring_procedente)

        #Escrevendo na planilha
        ws.cell(column=13, row=index, value="Sim")
        ws.cell(column=14, row=index, value=substring_procedente)
        ws.cell(column=15, row=index, value=data_ajustada[:11])
   
    #Se não houver as condições a cima lida com o erro escrevendo 'não há sentença' e escreve na planilha
    except NoSuchElementException as e:
        print('Não há Sentença')
        ws.cell(column=13, row=index, value="Não")
        ws.cell(column=14, row=index, value='-')
        ws.cell(column=15, row=index, value='-')

    #Aqui ocorre o tratamento e identificação do Recurso Ordinario    
    try:
        if driver.find_element(By.XPATH,'//*[contains(text(),"Recebido(s) o(s) Recurso Ordi")]'):
            recurso_ordinario = driver.find_element(By.XPATH,'//*[contains(text(),"Recebido(s) o(s) Recurso Ordi")]')
            data_recurso_ordinario = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(recurso_ordinario)).text
            data_ajustada = ajuste_data(data_recurso_ordinario)
            print('Data do Recurso:',str(data_ajustada)[:11])
           
            #Tratando a string do recurso
            start = recurso_ordinario.text.find('Recebido(s) o(s) Recurso Ordinário de ') + 38
            end = recurso_ordinario.text.find('sem efeito suspensivo')
            recurso_ordinario_substring = (recurso_ordinario.text[start:end])
            print('Recorrente (Recurso Ordinário): ',recurso_ordinario_substring)

            ws.cell(column=16, row=index, value='Sim')
            ws.cell(column=17, row=index, value=recurso_ordinario_substring)
            ws.cell(column=18, row=index, value=str(data_ajustada)[:11])
    except Exception as e:
            print("Não há Recurso Ordinário")
            ws.cell(column=16, row=index, value='Não')
            ws.cell(column=17, row=index, value='-')
            ws.cell(column=18, row=index, value='-')
            pass

    #Aqui ocorre o tratamento e identificação do Acórdão  
    try:
        if driver.find_element(By.XPATH,'//*[contains(text(),"(Acórdão)")]'):
            acordao = driver.find_element(By.XPATH,'//*[contains(text(),"(Acórdão)")]')
            acordao.click()
            data_acordao = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(acordao)).text
            data_ajustada = ajuste_data(data_acordao)
            print('Data do Acórdão:',str(data_ajustada)[:11])

            ws.cell(column=19, row=index, value='Sim')
            ws.cell(column=20, row=index, value=str(data_ajustada)[:11])
            try:
                #Emenda
                #Tem que ter esse sleep abaixo, se não da erro na identificação do parágrafo
                sleep(2)
                emenda = driver.find_element(By.XPATH,'//*[@id="visualizacao-documento"]/div/mat-card[2]/mat-card-content/span/div[6]').text
                emenda.strip()      
                try: #Metodo 1
                    if len(emenda) < 5:
                        print('Emenda (Acórdão): Incerto, verificar manualmente')
                        ws.cell(column=21, row=index, value='Incerto, verificar manualmente')
                    else:    
                        print('Emenda (Acórdão):',emenda)
                        ws.cell(column=21, row=index, value= emenda)
                except:#Metodo 2
                    emenda = driver.find_element(By.XPATH,'//*[@id="visualizacao-documento"]/div/mat-card[2]/mat-card-content/span/div[24]/p[2]').text
                    if len(emenda) < 5:
                        print('Emenda (Acórdão): Incerto, verificar manualmente')
                        ws.cell(column=21, row=index, value='Incerto, verificar manualmente')
                    else:    
                        print('Emenda (Acórdão):',emenda)
                        ws.cell(column=21, row=index, value= emenda)
            
            except Exception as e:
                print('Emenda não encontrada')
                ws.cell(column=21, row=index, value='Não')
                pass

                try:
                    #Provimento metodo 1
                    sleep(1) #Tem que ter esse sleep, é necessário!
                    if 'https://pje.trt15.jus.br/consultaprocessual/' in driver.current_url:
                        try: 
                            conclusao_provimento = driver.find_element(By.XPATH,'//*[contains(text(),"CONCLUSÃO")]')
                            provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"decido") or contains(text(),"decido:")]/parent::p/parent::div').below(conclusao_provimento)).text
                            if provimento:    
                                print("Provimento (Acórdão): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                            else:
                                provimento = provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"decido") or contains(text(),"decido:")]/parent::*').below(conclusao_provimento)).text
                                print("Provimento (Acórdão): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                                    
                        except:
                            try:
                                provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"decido") or contains(text(),"decido:") or contains(text(),"DECIDO:")]/parent::p/parent::div'))
                                print("Provimento (Acórdão): ",provimento)
                                ws.cell(column=22, row=index, value= provimento)  
                            except:
                                provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"decido") or contains(text(),"decido:") or contains(text(),"DECIDO:")]/parent::*'))
                                print("Provimento (Acórdão): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                                pass 

                            pass
                except Exception as e:
                    try:    
                            #Provimento metodo 2
                            conclusao_provimento = driver.find_element(By.XPATH,'//*[contains(text(),"CONCLUSÃO")]')
                            provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"provimento") or contains(text(),"PROVIMENTO")]').below(conclusao_provimento)).text
                            print("Provimento (Acórdão): ",provimento)
                            ws.cell(column=22, row=index, value= provimento) 
                    except:
                            provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"provimento") or contains(text(),"PROVIMENTO")]')).text
                            print("Provimento (Acórdão): ",provimento)
                            ws.cell(column=22, row=index, value= provimento) 

            try:
                #Conclusao 
                conclusao = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"ACORDAM") or contains(text(),"Acordam") or contains(text(),"A C O R D A M")]/.')).text
                print("Conclusão (Acórdão): ",conclusao)
                ws.cell(column=23, row=index, value= conclusao)  
                
            except:
                print("Conclusão (Acórdão): Incerto, verificar manualmente")
                ws.cell(column=23, row=index, value='Incerto, verificar manualmente')

    except Exception as e:
            print("Não há Acórdão em 1ª Instância")
            ws.cell(column=19, row=index, value='Não na 1ª Instância')
            ws.cell(column=20, row=index, value='-')
            ws.cell(column=21, row=index, value='-')
            ws.cell(column=22, row=index, value='-')
            ws.cell(column=23, row=index, value='-')
            pass

    # Clicando no cabeçalho para abrir o menu de infos do processo:
    WebDriverWait(driver, 3).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="titulo-detalhes"]/h1')))
    driver.find_element(By.XPATH, '//*[@id="titulo-detalhes"]/h1').click()

    # Coletando Orgão:
    orgao = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[1]').text
    ws.cell(column=2, row=index, value=orgao)

    # Coletando Distribuido:
    distribuido = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[2]').text
    ws.cell(column=3, row=index, value=distribuido)

    # Coletando Autuado:
    autuado = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[3]').text
    ws.cell(column=4, row=index, value=autuado)

    # Coletando Valor:
    valor = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dd[4]').text
    ws.cell(column=5, row=index, value=valor)

    # Coletando Justica:
    justica = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]/dl/dt[5]').text
    ws.cell(column=6, row=index, value=justica)

    #Check se ta vazia a celula G na linha = variavel index
    if ws[f'G{index}'].value == None:
        # Coletando Polo Ativo(Reclamante/Autor):
        polo_ativo = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[1]/pje-parte-processo/section/ul/li/span[2]').text
        ws.cell(column=7, row=index, value=polo_ativo)
        # Coletando Polo Passivo(Reclamado/Réu):
        polo_passivo = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[2]/pje-parte-processo/section/ul/li/span[2]').text
        ws.cell(column=9, row=index, value=polo_passivo)

    # Coletando Polo Ativo ADVOGADO:
    polo_ativo_adv = driver.find_element(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[1]/pje-parte-processo/section/ul/ul/li/small/span').text
    polo_ativo_adv.upper()
    ws.cell(column=8, row=index, value=polo_ativo_adv.removesuffix(" (ADVOGADO)"))


    # Coletando Polo Passivo ADVOGADO:
    try:
        lista_adv_web = []
        polo_passivo_adv = driver.find_elements(By.XPATH,'/html/body/pje-root/main/pje-detalhe-processo/div[6]/div[2]/div[2]/div[2]/pje-parte-processo/section/ul/ul/li')
        for value in polo_passivo_adv: 
            valor = str(value.text).upper().replace(' (ADVOGADO)','').strip()
            lista_adv_web.append(valor)
       
        if set(lista_adv_web).intersection(lista_proibida_adv):
            print("\n Atenção: ADVOGADO EM LISTA PROIBIDA, PRÓXIMO PROCESSO:",set(lista_adv_web).intersection(lista_proibida_adv))

            ws.cell(column=2, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=3, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=4, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=5, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=6, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=7, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=8, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=9, row=index, value='ADVOGADO EM LISTA PROIBIDA')
            ws.cell(column=10, row=index, value=f"ADVOGADO EM LISTA PROIBIDA {set(lista_adv_web).intersection(lista_proibida_adv)}")
            index += 1
            continue
        else:
            ws.cell(column=10, row=index, value=' , '.join(lista_adv_web).upper()) #Tem que usar esse join, se não da erro pra escrever a lista
    except Exception as e:
        ws.cell(column=10, row=index, value="Não disponível")
        print(e)

    # Coletando Assuntos:
    sleep(1)
    orgao = driver.find_element(By.XPATH, '//*[@id="colunas-dados-processo"]/div[1]').text +'...'
    x = str(orgao)  # Garantindo que ta em formato de string
    inicio = x.find("Assunto(s):") + 12  # Encontrando o elemento "Assuntos" e definindo que será o inicio da substring, adicionar +11 exclui o elemento (Assunto(s) tem 11 chars)
    final = x.find("...",inicio)  # Note que na linha 123 adicionei reticencias para delimitar como sendo o final da string orgao
    substring = x[inicio:final]  # Criando a substring, que vem da 'x' a partir do inicio até o final
    ws.cell(column=12, row=index, value=substring)  # Inserindo na planilha
    
    # Salvando a planilha com os dados
    wb.save(file)
    print("\nSalvando..")
    sleep(0.5)

    #COLETA DO ACÓRDÃO NA SEGUNDA INSTANCIA
    #_____________________________________#
    if ws.cell(row=index,column=19).value == "Sim": #Se encontrar o Acórdão no 1grau, faz essa linha
        try:
            print(f'\nConsulta {index - 1} finalizada..')
            os.remove(imagem)
            driver.quit()

            timex = round((time() - start_time))
            if timex < 60:
                print('Tempo da consulta: %.1f segundos' % (timex))
            elif timex < 3600:
                print('Tempo da consulta: %.1f minuto(s)' % (timex / 60))
            else:
                print('Tempo da consulta: %.1f hora(s)' % (timex / 3600))
            # Add +1 a variavel global index, para que ela seja '2' no proximo loop e leia a próxima linha
            index += 1
            continue
        except PermissionError:
            print("\nATENÇÃO: Planilha aberta, fechar a mesma e reiniciar a automação")
            break

    if ws.cell(row=index,column=19).value == "Não" or "Não na 1ª Instância": #Verifica se deu "Não" em 1grau, se sim vai pro 2grau
                
        driver.get(f'https://pje.trt{numero_trt}.jus.br/consultaprocessual/')
        print("\nRealizando a consulta de 2ª Instância no processo: " + str(num_processo))   
        # Pegando o número do processo do excel e entrando na pagina
        WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.ID, 'label-sistema')))
        driver.find_element(By.XPATH, '//*[@id="nrProcessoInput"]').send_keys(num_processo)
        driver.find_element(By.XPATH, '//*[@id="btnPesquisar"]').send_keys(Keys.ENTER)
        
        # Clicando no botao do 2Grau
        sleep(1)
        try:
            #Há processos que só tem 1grau e 2grau, outros pulam diretamente para a tela do processo
            driver.find_element(By.XPATH,'//*[@id="painel-escolha-processo"]/button[2]').click()

        except:
            pass #isso serve pra continuar se ele não encontrar nenhum dos dois botões

        y=0
        while y < 1: #Esse loop é para ele repetir a tentativa de escrever a captcha
            try:
                # Escrevendo a Captcha
                WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.XPATH,'//*[@id="btnEnviar"]')))
                sleep(1)
                driver.find_element(By.XPATH,'//*[@id="imagemCaptcha"]').screenshot(r'C:\Users\abcd\Desktop\captcha.png')
                #Local onde salvar o print
                imagem = (r'C:\Users\abcd\Desktop\captcha.png')

                #Pega a imagem e envia para o servidor 2captcha
                result = solver.normal(imagem) #'Result' trás 2 resultados: 'captchaId' e 'code', queremos apenas o 'code'
                print("Captcha:",result['code'])

                driver.find_element(By.XPATH,'//*[@id="captchaInput"]').send_keys(result['code'])
                driver.find_element(By.XPATH, '//*[@id="btnEnviar"]').click()
                y+=1
            
            #Aqui até o continue faz com que atualize a pagina caso dê erro na API da captcha
            except Exception as e:
                print("Erro na Captcha:",e)
                driver.refresh()
                sleep(2)
                pass

        #Aguarda o emblema do Brasil  no documento para poder prosseguir
        WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.XPATH, '//*[@id="visualizacao-documento"]/div/mat-card/mat-card-header/div[2]')))
        print('Iniciando coleta de dados do processo em 2ª Instancia...')
        #Aqui ocorre o tratamento e identificação do Acórdão  
        try:
            if driver.find_element(By.XPATH,'//*[contains(text(),"(Acórdão)")]'):
                acordao = driver.find_element(By.XPATH,'//*[contains(text(),"(Acórdão)")]')
                acordao.click()
                data_acordao = driver.find_element(locate_with(By.XPATH,"//*[contains(@name,'dataItemTimeline')]").above(acordao)).text
                data_ajustada = ajuste_data(data_acordao)
                print('Data do Acórdão 2ª Instância:',str(data_ajustada)[:11])

                ws.cell(column=19, row=index, value='Sim, em 2ª Instância')
                ws.cell(column=20, row=index, value=str(data_ajustada)[:11])
                try:
                    #Emenda
                    #Tem que ter esse sleep abaixo, se não da erro na identificação do parágrafo
                    sleep(2)
                    emenda = driver.find_element(By.XPATH,'//*[@id="visualizacao-documento"]/div/mat-card[2]/mat-card-content/span/div[6]').text
                    emenda.strip()      
                    try: #Metodo 1
                        if len(emenda) < 5:
                            print('Emenda (Acórdão 2ª Instância): Incerto, verificar manualmente na 2ª Instancia')
                            if ws.cell(column=21,row=index).value == NoneType or '-' or 'Não':
                                ws.cell(column=21, row=index, value='Incerto, verificar manualmente na 2ª Instancia')
                        else:    
                            print('Emenda (Acórdão 2ª Instância):',emenda)
                            if ws.cell(column=21,row=index).value == NoneType or '-' or 'Não':
                                ws.cell(column=21, row=index, value= emenda)
                    except:#Metodo 2
                        emenda = driver.find_element(By.XPATH,'//*[@id="visualizacao-documento"]/div/mat-card[2]/mat-card-content/span/div[24]/p[2]').text
                        if len(emenda) < 5:
                            print('Emenda (Acórdão 2ª Instância): Incerto, verificar manualmente na 2ª Instancia')
                            ws.cell(column=21, row=index, value='Incerto, verificar manualmente na 2ª Instancia')
                        else:    
                            print('Emenda (Acórdão 2ª Instância):',emenda)
                            if ws.cell(column=21,row=index).value == NoneType or '-' or 'Não':
                                ws.cell(column=21, row=index, value= emenda)
                
                except Exception as e:
                    print('Emenda não encontrada 2ª Instância')
                    ws.cell(column=21, row=index, value='Não')
                    pass

                try:
                    #Provimento metodo 1
                    sleep(1) #Tem que ter esse sleep, é necessário!
                    provimento = driver.find_element(By.XPATH,"//*[contains(text(),'embargos de declaração') or contains(text(),'EMBARGOS DE DECLARAÇÃO')]/parent::*").text
                    print('\nProvimento (Acórdão 2ª Instância):',provimento)
                    if ws.cell(column=22,row=index).value == None or '-' or 'Não':
                        ws.cell(column=22, row=index, value= provimento)
                    
                    if 'https://pje.trt15.jus.br/consultaprocessual/' in driver.current_url:
                        try: 
                            provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"embargos de declaração")]').above(acordao)).text
                            if provimento:    
                                print("Provimento (Acórdão 2ª Instância): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                            else:
                                provimento = provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"embargos de declaração")]').above(acordao)).text
                                print("Provimento (Acórdão 2ª Instância): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                                    
                        except:
                            try:
                                provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"embargos de declaração")]')).text
                                print("Provimento (Acórdão 2ª Instância): ",provimento)
                                ws.cell(column=22, row=index, value= provimento)  
                            except:
                                provimento = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"embargos de declaração")]')).text
                                print("Provimento (Acórdão 2ª Instância): ",provimento)
                                ws.cell(column=22, row=index, value= provimento) 
                                pass 
                    
                except Exception as e:
                        try: 
                            provimento = driver.find_element(By.XPATH,'//*[contains(text(),"decido") or contains(text(),"decido:")]').text
                            print("Provimento (Acórdão 2ª Instância): ",provimento)
                            if ws.cell(column=22,row=index).value == None or '-' or 'Não':
                                    ws.cell(column=22, row=index, value= provimento)          
                        except Exception as e:
                            pass


            try:
                #Conclusao 
                conclusao = driver.find_element(locate_with(By.XPATH,'//*[contains(text(),"ACORDAM") or contains(text(),"Acordam") or contains(text(),"JULGOU") or contains(text(),"A C O R D A M")]/.')).text
                print("Conclusão (Acórdão 2ª Instância): ",conclusao)
                ws.cell(column=23, row=index, value= conclusao)  
                    
            except:
                print("Conclusão (Acórdão 2ª Instância): Incerto, verificar manualmente")
                ws.cell(column=23, row=index, value='Incerto, verificar manualmente na 2ª Instancia')

        except Exception as e:
                print("Não há Acórdão na 2ª Instância",e)
                ws.cell(column=19, row=index, value='Não na 1ª nem 2ª Instância')
                ws.cell(column=20, row=index, value='-')
                ws.cell(column=21, row=index, value='-')
                ws.cell(column=22, row=index, value='-')
                ws.cell(column=23, row=index, value='-')

        try:
            # Salvando a planilha com os dados
            wb.save(file)
            print(f'\nConsulta {index - 1} finalizada..')
            os.remove(imagem)
            driver.quit()

            timex = round((time() - start_time))
            if timex < 60:
                print('Tempo da consulta: %.1f segundos' % (timex))
            elif timex < 3600:
                print('Tempo da consulta: %.1f minuto(s)' % (timex / 60))
            else:
                print('Tempo da consulta: %.1f hora(s)' % (timex / 3600))
            # Add +1 a variavel global index, para que ela seja '2' no proximo loop e leia a próxima linha
            index += 1
            continue
            
        except PermissionError:
            print("\nATENÇÃO: Planilha aberta, fechar a mesma e reiniciar a automação")
            break

tempo2 = round((time() - start_time))
if tempo2 < 60:
    print('Programa finalizado | Tempo Total: %.1f segundos' % (tempo2))
elif tempo2 < 3600:
    print('Programa finalizado | Tempo Total: %.1f minuto(s)' % (tempo2 / 60))
else:
    print('Programa finalizado | Tempo Total: %.1f hora(s)' % (tempo2 / 3600))

print('Saldo 2Captcha:',solver.balance(),'$USD')
